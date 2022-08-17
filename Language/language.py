# -*- coding: utf-8 -*-
# @Time : 2022/3/28 11:46
# @Author : zengf
# @Email : 944458157@qq.com
# @File : language.py
# @Project : Language

import xlrd
from openpyxl import load_workbook
import wx
import sys
import csv
from zhconv import convert
import threading


class Language():
    def __init__(self, oldpath, newpath, project):
        self.oldpath = oldpath
        self.newpath = newpath
        self.project = project
        self.old = None
        self.new = None
        self.country = []
        self.newids = []
        self.add = []

    def read_file(self):
        old = xlrd.open_workbook(self.oldpath)
        new = xlrd.open_workbook(self.newpath)
        if self.project in ["PunBall", "Kinja", "Survivor"]:
            self.old = old.sheet_by_name('Language')
            self.new = new.sheet_by_name('Language')
        if self.project in ["Archero"]:
            self.old = old.sheet_by_name('lauguage')
            self.new = new.sheet_by_name('lauguage')
        if self.project in ["Slidey"]:
            self.old = old.sheet_by_name('Sheet1')
            self.new = new.sheet_by_name('Sheet1')
        if self.project == "Archero":
            self.country = self.new.row_values(2, start_colx=1)  # 获取多语言国家种类列表
        if self.project == "PunBall" or self.project == "Survivor":
            self.country = self.new.row_values(2, start_colx=3)  # 获取多语言国家种类列表
        if self.project == "Kinja":
            self.country = self.new.row_values(0, start_colx=4)  # 获取多语言国家种类列表
        if self.project == "Slidey":
            self.country = self.new.row_values(0, start_colx=1)  # 获取多语言国家种类列表
        self.olddic = {}
        self.newdic = {}
        if self.project == "Archero":
            self.oldids = self.old.col_values(0)  # 获取多语言id
            self.newids = self.new.col_values(0)  # 获取多语言id
            self.oldtip = self.old.col_values(1)  # 获取简体中文描述列表
            self.newtip = self.new.col_values(1)  # 获取简体中文描述列表
        if self.project == "PunBall":
            self.oldids = self.old.col_values(1)  # 获取多语言id
            self.newids = self.new.col_values(1)  # 获取多语言id
            self.oldtip = self.old.col_values(3)  # 获取简体中文描述列表
            self.newtip = self.new.col_values(3)  # 获取简体中文描述列表
        if self.project == "Kinja":
            self.oldids = self.old.col_values(1)  # 获取多语言id
            self.newids = self.new.col_values(1)  # 获取多语言id
            self.oldtip = self.old.col_values(4)  # 获取简体中文描述列表
            self.newtip = self.new.col_values(4)  # 获取简体中文描述列表
        if self.project == "Slidey":
            self.oldids = self.old.col_values(0)  # 获取多语言id
            self.newids = self.new.col_values(0)  # 获取多语言id
            self.oldtip = self.old.col_values(5)  # 获取简体中文描述列表
            self.newtip = self.new.col_values(5)  # 获取简体中文描述列表
        if self.project == "Survivor":
            self.oldids = self.old.col_values(0)  # 获取多语言id
            self.newids = self.new.col_values(0)  # 获取多语言id
            self.oldtip = self.old.col_values(3)  # 获取简体中文描述列表
            self.newtip = self.new.col_values(3)  # 获取简体中文描述列表
        for i in range(len(self.oldids)):
            self.olddic[self.oldids[i]] = self.oldtip[i]
        for j in range(len(self.newids)):
            self.newdic[self.newids[j]] = self.newtip[j]

        lose = list(set(self.oldids).difference(set(self.newids)))
        # 检查新版本的language对比上个版本有没有缺失
        if lose != []:
            print("新版本丢失的多语言：{0}".format(lose) + "\n")
        else:
            print("*****新版本没有丢失多语言*****\n")
        temp = list(set(self.oldids) - set(lose))
        self.add = list(set(self.newids).difference(set(temp)))  # 获取新增的多语言id列表

        ch = 0
        for l in temp:
            if self.olddic[l] != self.newdic[l]:
                ch += 1
                self.add.append(l)

        print("*****新版本修改了{0}".format(ch) + "条多语言*****")
        if ch > 0:
            for i in self.add[-ch:]:
                print(self.olddic[i])
        print()
        print("*****新版本新增了{0}".format(len(self.add)-ch)+"条多语言*****\n")

    def check_country_null(self):
        try:
            csvFile2 = open('LanguageDiff.csv', 'w', newline='', encoding='utf-8')  # 设置newline，否则两行之间会空一行
            # csvFile2.truncate()
            writer = csv.writer(csvFile2)
            writer.writerow(['多语言', '亚洲最长语言', '亚洲最长', '其他最长语言', '其他最长'])
            for i in range(len(self.add)):
                ind = self.newids.index(self.add[i])
                if self.project == "Archero":
                    countrys = self.new.row_values(ind, start_colx=1)
                    asia = countrys[0:2] + countrys[7:9]  # 亚洲国家
                    other = countrys[2:7] + countrys[9:]  # 非亚洲国家
                if self.project == "PunBall":
                    countrys = self.new.row_values(ind, start_colx=3)
                    asia = countrys[2:4] + countrys[0:1] + countrys[8:9]  # 亚洲国家
                if self.project == "Kinja":
                    countrys = self.new.row_values(ind, start_colx=4)
                    asia = countrys[2:4] + countrys[0:1] + countrys[14:15]  # 亚洲国家
                    other = countrys[4:14] + countrys[1:2] + countrys[15:]  # 非亚洲国家
                if self.project == "Slidey":
                    countrys = self.new.row_values(ind, start_colx=1)
                    asia = countrys[4:6] + countrys[8:10]  # 亚洲国家
                    other = countrys[0:4] + countrys[6:8] + countrys[10:]  # 非亚洲国家
                if self.project == "Survivor":
                    countrys = self.new.row_values(ind, start_colx=3)
                    asia = countrys[0:1] + countrys[6:8] + countrys[15:]  # 亚洲国家
                    other = countrys[1:6] + countrys[8:15]  # 非亚洲国家
                    for a in range(len(asia)):
                        asia[a] = str(asia[a])
                    for o in range(len(other)):
                        other[o] = str(other[o])
                # 检查新增的多语言有哪些没有配置全
                if '' in countrys:
                    if self.project == "Slidey":
                        countrys.pop(-1)
                        if '' in countrys:
                            writer.writerow([countrys[4], '多语言配置不全!!!!!'])
                        else:
                            asiaLongest = max(asia, key=len, default='')
                            otherLongest = max(other, key=len, default='')
                            asial = self.country[countrys.index(asiaLongest)]
                            otherl = self.country[countrys.index(otherLongest)]
                            writer.writerow([countrys[4], asial, asiaLongest, otherl, otherLongest])
                    elif self.project == "Archero":
                        countrys = countrys[0:16]
                        if '' in countrys:
                            writer.writerow([countrys[0], '多语言配置不全!!!!!'])
                        else:
                            asiaLongest = max(asia, key=len, default='')
                            otherLongest = max(other, key=len, default='')
                            asial = self.country[countrys.index(asiaLongest)]
                            otherl = self.country[countrys.index(otherLongest)]
                            writer.writerow([countrys[0], asial, asiaLongest, otherl, otherLongest])
                    else:
                        writer.writerow([countrys[0], '多语言配置不全!!!!!'])
                else:
                    asiaLongest = max(asia, key=len, default='')
                    otherLongest = max(other, key=len, default='')
                    asial = self.country[countrys.index(asiaLongest)]
                    otherl = self.country[countrys.index(otherLongest)]
                    writer.writerow([countrys[0], asial, asiaLongest, otherl, otherLongest])

            csvFile2.close()
        except Exception as e:
            raise Exception(e)
            print("写入文件失败！！！可能是文件正打开着，需要把文件先关闭~~")


class Waibao():
    def __init__(self, oldpath, newpath, project="Survivor"):
        self.waibaoPath = oldpath
        self.survivorPath = newpath
        self.project = project
        self.old = None
        self.new = None
        self.country = []
        self.newids = []
        self.add = []
        self.check = True

    def read_file(self):
        epiboly = xlrd.open_workbook(self.waibaoPath)
        survivor = xlrd.open_workbook(self.survivorPath)
        self.sur = survivor.sheet_by_name('Language')
        self.country = self.sur.row_values(2, start_colx=3)  # 获取多语言国家种类列表
        self.surids = self.sur.col_values(0, start_rowx=3)  # 获取多语言id
        self.surtip = self.sur.col_values(3, start_rowx=3)  # 获取简体中文描述列表
        self.surdic = dict(map(lambda x, y: [x, y], self.surids, self.surtip))

        self.epi = epiboly.sheet_by_name('Sheet1')
        # self.epidic = {}
        self.epiids = self.epi.col_values(0, start_rowx=2)  # 获取多语言id
        self.epiCh = self.epi.col_values(1, start_rowx=2)  # 获取简体中文描述列表
        self.epidic = dict(map(lambda x, y: [x, y], self.epiids, self.epiCh))

        self.otherLan = {}
        for s in range(len(self.epiids)):
            self.otherLan[self.epiids[s]] = self.epi.row_values(s+2, start_colx=4, end_colx=19)

        print("*****1.检查外包翻译表中是否缺少id或中文描述*****\n")
        for id in range(len(self.epiids)):
            if not self.epiids[id]:
                print("×××第{0}行缺少多语言id".format(id+3))
                self.check = False
        for c in range(len(self.epiCh)):
            if not self.epiCh[c]:
                print("×××第{0}行缺少多语言中文".format(c+3))
                self.check = False
        if self.check:
            print("*****2.检查外包翻译表和多语言表能否匹配上*****\n")
            for k in range(len(self.epiids)):
                if self.epiids[k] and self.epiids[k] not in self.surids:
                    print("×××外包翻译中的多语言id{0}在项目的多语言表中未找到".format(self.epiids[k]))
                    self.check = False
                if self.epiids[k] in self.surids and self.surdic[self.epiids[k]] != self.epiCh[k]:
                    print(self.epiCh[k], type(self.epiCh[k]))
                    print(self.surdic[self.epiids[k]], type(self.surdic[self.epiids[k]]))
                    print("×××外包翻译中多语言id{0}对应的中文与项目多语言表中的中文对不上".format(self.epiids[k]))
                    self.check = False
                if convert(self.epiCh[k], 'zh-tw') != self.otherLan[self.epiids[k]][-1]:
                    print("警告（不影响导入）外包翻译中多语言id{0}工具翻译的繁中与它的对不上".format(self.epiids[k]))
                    print(convert(self.epiCh[k], 'zh-tw'), self.otherLan[self.epiids[k]][-1])
        if self.check:
            print("*****3.检查外包翻译表多语言是否齐全*****\n")
            for k, v in self.otherLan.items():
                if '' in v:
                    print("×××外包翻译中多语言id{0}的翻译不全".format(k))
                    self.check = False

    def start_import(self):
        try:
            print('*****正在写入*****\n')
            wb = load_workbook(self.survivorPath)
            sheet = wb.worksheets[0]
            for i in self.epiids:
                colx = self.surids.index(i)
                for l in range(len(self.otherLan[i])):
                    sheet.cell(colx+4, l+5).value = self.otherLan[i][l]
            wb.save(self.survivorPath)
            print('*****文件写入完毕*****\n')
        except Exception as e:
            print("写入文件失败！！！可能是文件正打开着，需要把文件先关闭~~")
            raise Exception(e)


class checkThread(threading.Thread):
    def __init__(self, devs):
        threading.Thread.__init__(self)
        self.devs = devs

    def run(self):
        self.devs.start_import()


class RedirectText(object):
    def __init__(self, aWxTextCtrl):
        self.out = aWxTextCtrl

    def write(self, string):
        self.out.WriteText(string)


class MyApp(wx.App):
    def OnInit(self):
        mm = wx.DisplaySize()
        self.frame = wx.Frame(parent=None, title="多语言辅助工具", size=(mm[0], mm[1]))
        panel = wx.Panel(self.frame, -1)
        self.label1 = wx.StaticText(panel, -1, "旧版本language.xls", pos=(10, 10))
        self.text1 = wx.TextCtrl(panel, -1, pos=(150, 10), size=(250, 20))
        self.choose1 = wx.Button(panel, -1, '选择文件', pos=(450, 10))
        self.waibao = wx.StaticText(panel, -1, "外包回来的翻译", pos=(650, 10))
        self.waibaopath = wx.TextCtrl(panel, -1, pos=(770, 10), size=(250, 20))
        self.choose3 = wx.Button(panel, -1, '选择文件', pos=(1070, 10))
        self.label2 = wx.StaticText(panel, -1, "新版本language.xls", pos=(10, 60))
        self.text2 = wx.TextCtrl(panel, -1, pos=(150, 60), size=(250, 20))
        self.choose2 = wx.Button(panel, -1, '选择文件', pos=(450, 60))
        self.totle = wx.StaticText(panel, -1, "多语言总表", pos=(650, 60))
        self.totlepath = wx.TextCtrl(panel, -1, pos=(770, 60), size=(250, 20))
        self.choose4 = wx.Button(panel, -1, '选择文件', pos=(1070, 60))
        self.log = wx.TextCtrl(panel, wx.ID_ANY, pos=(20, 200), size=(1500, 600),
                          style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)
        self.label3 = wx.StaticText(panel, -1, "选择项目", pos=(10, 110))
        self.project = wx.ComboBox(panel,  -1, value="请选择", choices=["Archero", "PunBall", "Kinja", "Slidey", "Survivor"], pos=(70, 110))
        self.button = wx.Button(panel, -1, '检查', pos=(250, 110))
        self.imp = wx.Button(panel, -1, '导入', pos=(1070, 110))

        # 按钮绑定事件
        self.Bind(wx.EVT_BUTTON, self.check, self.button)
        self.Bind(wx.EVT_BUTTON, self.lanImport, self.imp)
        self.Bind(wx.EVT_BUTTON, self.chooseOldFile, self.choose1)
        self.Bind(wx.EVT_BUTTON, self.chooseNewFile, self.choose2)
        self.Bind(wx.EVT_BUTTON, self.chooseOldFileL, self.choose3)
        self.Bind(wx.EVT_BUTTON, self.chooseNewFileL, self.choose4)

        # redirect text here
        redir = RedirectText(self.log)
        sys.stdout = redir

        self.frame.Show()
        return True

    def chooseOldFile(self, event):
        dlg = wx.FileDialog(self.frame, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.text1.SetValue(dlg.GetPath())
        dlg.Destroy()

    def chooseNewFile(self, event):
        dlg = wx.FileDialog(self.frame, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.text2.SetValue(dlg.GetPath())
        dlg.Destroy()

    def chooseOldFileL(self, event):
        dlg = wx.FileDialog(self.frame, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.waibaopath.SetValue(dlg.GetPath())
        dlg.Destroy()

    def chooseNewFileL(self, event):
        dlg = wx.FileDialog(self.frame, u"选择文件夹", style=wx.DD_DEFAULT_STYLE)
        if dlg.ShowModal() == wx.ID_OK:
            self.totlepath.SetValue(dlg.GetPath())
        dlg.Destroy()

    def check(self, event):
        oldfile = self.text1.GetValue()
        newfile = self.text2.GetValue()
        pro = self.project.GetValue()
        if pro == "请选择":
            print("没有选择项目！！！！！")
        else:
            lan = Language(oldfile, newfile, self.project.GetValue())
            lan.read_file()
            lan.check_country_null()
            print('*****文件写入完毕*****\n')

    def lanImport(self, event):
        oldfile = self.waibaopath.GetValue()
        newfile = self.totlepath.GetValue()
        print('*****开始检查，请等待几秒*****\n')
        surviorIo = Waibao(oldfile, newfile)
        surviorIo.read_file()
        if not surviorIo.check:
            print('*****请确认好文件的问题再导入！*****\n')
        else:
            run = checkThread(surviorIo)
            run.start()

    def clear(self, event):
        self.log.Clear()


if __name__ == '__main__':
    app = MyApp()  # 启动
    app.MainLoop()  # 进入消息循环
