# -*- coding: utf-8 -*-
# @Path : C:\Users\zengf\AppData\Local\Programs\Python\Python39
# @Time : 2022/5/25 14:20
# @Author : zengf
# @Email : 944458157@qq.com
# @File : logtoexcel.py
# @Project : workspace
import os
from openpyxl import load_workbook
import xlrd

BattlPath = 'C:\\Users\\zengf\\AppData\\LocalLow\\habby\\PunBall\\Damage\\20220609'
ToolPath = "D:\\fanwei\\弹球\\数值公式\\punball伤害公式_v4.xlsx"
SkillXls = "D:\\fanwei\\弹球\\数值公式\\Skill.xls"
Monster = {"100001":"骷髅小兵","100024":"蜘蛛","100060":"火元素","100059":"冰元素","100006":"愤怒怪","100002":"炸弹怪",
           "100004":"毒罐","100025":"弩车","100079":"泥沙怪","100020":"牧师","100003":"盾兵","100021":"召唤怪",
           "100015":"分裂史莱姆","100073":"忍者","100078":"双管豌豆","100018":"幽灵怪","100019":"石像鬼","100005":"冲锋猪",
           "100066":"吸血鬼","100077":"弓箭手","100017":"豌豆射手","100082":"章鱼怪","100083":"圣骑士","100081":"哥布林",
           "600001":"年兽-夕","100084":"爆竹","600002":"海德拉","100086":"海德拉之大头","100085":"海德拉之小头",
           "100087":"光元素","100088":"暗元素","300006":"精英蜘蛛","300015":"精英火元素","300016":"精英冰元素",
           "110006":"精英愤怒怪","300033":"精英炸弹怪","300034":"精英毒罐","300010":"精英弩车","300031":"精英泥沙怪",
           "300004":"精英牧师","110003":"精英盾兵","300003":"精英召唤怪","300002":"精英分裂史莱姆","110025":"精英忍者",
           "110078":"精英双管豌豆","110018":"精英幽灵怪","110019":"精英石像鬼","110005":"精英冲锋猪","300032":"精英吸血鬼",
           "110077":"精英弓箭手","110017":"精英豌豆射手","110082":"精英章鱼怪","110083":"精英圣骑士","110084":"精英光元素",
           "110085":"精英暗元素","500003":"Boss蜘蛛","500018":"Boss火元素","500019":"Boss冰元素","500020":"Boss愤怒怪",
           "500021":"Boss弩车","500025":"Boss泥沙怪","500005":"Boss牧师","500008":"Boss盾兵","500004":"Boss召唤怪",
           "500002":"Boss分裂史莱姆","500009":"Boss忍者","500022":"Boss双管豌豆","500010":"Boss幽灵怪","500011":"Boss石像鬼",
           "500023":"Boss冲锋猪","500012":"Boss吸血鬼","500024":"Boss弓箭手","500001":"Boss豌豆射手","500029":"Boss章鱼怪",
           "500030":"Boss圣骑士","500031":"Boss光元素","500032":"Boss暗元素"}
BallType = {"Normal":"普通球","Fire":"火球","Poison":"毒球","Thunder":"雷球","BlackHole":"黑洞球","Bomb":"炸弹球",
            "Penetrate":"穿透球","Split":"分裂球","AddBall":"添加新球","Ice":"冰球","Laser":"激光球","Snipe":"狙击球",
            "Melee":"近战球","BackStab":"背刺球","Missile":"飞弹球","SecKill":"秒杀球","Holy":"神圣球"}
SkillType = {"Normal":"正常受击","Critical10":"10倍暴击","Critical100":"100倍暴击","SecKill":"秒杀","Block":"格挡",
             "Miss":"闪避","Immune":"免疫","Invincible":"无敌","BackStab":"背刺","BombMonster":"灼炸弹怪","Bomb":"炸弹",
             "Fire":"火","Ice":"冰","Thunder":"雷","Laser":"光","BlackHole":"黑洞","Missile":"飞弹","Poison":"毒",
             "Holy":"神圣","Absorb":"吸收"}
Total = [["描述", "类型", "怪物类型", "伤害类型", "全技能", "列表", "可能生效", "列表", "基础伤害", "X", "x1", "x2", "B", "b1",
          "C", "c1", "c2", "c3", "c4", "M", "D", "d1", "d2", "d3", "d4", "E", "最终伤害"]]
SkillData = [["基础伤害", "X", "x1", "x2", "B", "b1", "C", "c1", "c2", "c3", "c4", "c5", "c6", "c7", "c8", "c9", "c10", "M",
              "D", "d1", "d2", "d3", "d4", "d5", "d6", "d7", "d8", "d9", "d10", "E", "最终伤害"]]
SkillSkill = []
BallToSkill = {
	"普通球" : ["400047", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"火球" : ["400048", "400023", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"毒球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"雷球" : ["400048", "400025", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"黑洞球" : ["400048", "400027", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"炸弹球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"穿透球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"分裂球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"添加新球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"冰球" : ["400048", "400024", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"激光球" : ["400048", "400026", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"狙击球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"近战球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"背刺球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"飞弹球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"秒杀球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"神圣球" : ["400048", "521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"],
	"Player": ["521001", "523001", "528001", "521003", "523003", "528003", "531001", "533001", "538001", "541002", "543002", "548002", "541003", "543003", "548003", "551002", "553002", "558002", "551003", "553003", "558003", "561001", "563001", "568001", "561003", "563003", "568003", "561004" , "563004", "568004", "511004", "513004", "518004", "541004", "543004", "548004", "551004", "553004", "558004"]
}
First = ['400045']
Last = ['400046']
SkillList = []


def getlog(path):
    global Total, SkillList, SkillData, SkillSkill
    with open(path, encoding='utf-8 ') as file:
        content = file.readlines()
    for line in content:
        # print(line)
        round = []
        sk = []
        maybeskill = []
        line.rstrip()
        ratio = line.split("}；")
        for r in range(len(ratio)):
            if r == 0:
                dic = ratio[0].split("；攻击力修正系数：")[0].split("；")
                # print(dic)
                xdata = ""
                if "攻击力修正系数" in ratio[0]:
                    xdata = ratio[0].split("；攻击力修正系数：")[1]
                if "怪物受伤" in dic[0]:
                    for i in range(len(dic)):
                        target = dic[i].split("：")[-1]
                        if "PlayerSkill" not in dic[0] and "Buff" not in dic[0] and "FixedDamage" not in dic[0] and \
                                "PetSkill" not in dic[0] and "Seckill" not in dic[0]:
                            if 1 == i:
                                if "FirstBall_Ball" in dic[1]:
                                    round.append('FirstBall_'+BallType[target])
                                    continue
                                elif "LastBall_Ball" in dic[1]:
                                    round.append('LastBall_'+BallType[target])
                                    continue
                                else:
                                    round.append(BallType[target])
                                    continue
                        if 'Player102(Clone) (Punball.Game.Player.PlayerEntity)' == target:
                            round.append('Player')
                            continue
                        if '(Punball.MonsterBase)' in target:
                            mid = target.split(" ")[0][7:13]
                            if mid in Monster.keys():
                                round.append(Monster[mid])
                            else:
                                round.append(target)
                            continue
                        round.append(target)
                    sk.append(dic[-1].split("：")[-1])
                    if "PlayerSkill" in round[0] or "Buff" in round[0]:
                        maybeskill = list(set(maybeskill).union(set(BallToSkill["Player"]) & set(SkillList)))
                    if round[1] in BallToSkill.keys():
                        maybeskill = list(set(maybeskill).union(set(BallToSkill[round[1]]) & set(SkillList)))
                    if "_" in round[1]:
                        if round[1].split("_")[0] == "FirstBall":
                            maybeskill = list(set(maybeskill).union(set(BallToSkill[round[1].split("_")[1]]+First) & set(SkillList)))
                        if round[1].split("_")[0] == "LastBall":
                            maybeskill = list(set(maybeskill).union(set(BallToSkill[round[1].split("_")[1]]+Last) & set(SkillList)))
                elif "添加技能" in ratio[0]:
                    SkillList = ratio[0].split("：")[2].rstrip("；\n").split("；")
                    break
                else:
                    break
                round.append("所有技能：")
                round.append(str(SkillList))
                round.append("可能影响参数的技能：")
                round.append(str(maybeskill))
                if "攻击力增加" not in xdata:
                    X = xdata.split("{")[0]
                    round.append(X)
                    round.append('0')
                    round.append('0')
                    sk.append(X)
                    sk.append('0')
                    sk.append('0')
                else:
                    X = xdata.split("{")[0]
                    x1 = xdata.split("；")[0].split("：")[1]
                    round.append(X)
                    round.append(x1)
                    sk.append(X)
                    sk.append(x1)
                    if "球" in xdata:
                        x2 = xdata.split("；")[1].split("：")[1]
                        round.append(x2)
                        sk.append(x2)
                    else:
                        round.append("0")
                        sk.append("0")
            if r == 1:
                if "基础倍数" in ratio[r] or "传入系数" in ratio[r]:
                    B = ratio[r].split("：")[1].split("{")[0]
                    round.append(B)
                    sk.append(B)
                    if "球" not in ratio[r]:
                        b1 = ratio[r].split("：")[2]
                    else:
                        b1 = str(float(ratio[r].split("：")[3].split("；")[0]) * float(ratio[r].split("：")[2].split("；")[0]))
                    round.append(b1)
                    sk.append(b1)
                else:
                    B = ratio[r].split("：")[1].split("{")[0]
                    round.append(B)
                    round.append("1")
                    sk.append(B)
                    sk.append("1")
            if r == 2:
                if "属性" in ratio[r]:
                    Clist = ratio[r].split("；")
                    C = Clist[0].split("：")[1].split("{")[0]
                    round.append(C)
                    round.append(Clist[0].split("：")[2])
                    sk.append(C)
                    sk.append(Clist[0].split("：")[2])
                    cnum = len(Clist)
                    for i in range(cnum-2):
                        round.append(Clist[i+1].split("：")[1])
                        sk.append(Clist[i+1].split("：")[1])
                    for i in range(10-cnum+1):
                        round.append("0")
                        sk.append("0")
                else:
                    C = ratio[r].split("：")[1].split("{")[0]
                    round.append(C)
                    sk.append(C)
                    for i in range(10):
                        round.append("0")
                        sk.append("0")
            if r == 3:
                if "属性" in ratio[r]:
                    Dlist = ratio[r].split("；")
                    M = Dlist[0].split("：")[1]
                    D = Dlist[1].split("：")[1].split("{")[0]
                    d1 = Dlist[1].split("：")[2]
                    round.append(M)
                    round.append(D)
                    round.append(d1)
                    sk.append(M)
                    sk.append(D)
                    sk.append(d1)
                    dnum = len(Dlist)
                    for i in range(dnum-3):
                        round.append(Dlist[i+2].split("：")[1])
                        sk.append(Dlist[i+2].split("：")[1])
                    for i in range(10-dnum+1):
                        round.append("0")
                        sk.append("0")
                else:
                    M = ratio[r].split("：")[1].split("；")[0]
                    D = ratio[r].split("：")[2].split("{")[0]
                    round.append(M)
                    round.append(D)
                    sk.append(M)
                    sk.append(D)
                    for i in range(10):
                        round.append("0")
                        sk.append("0")
            if r == 4:
                round.append(ratio[r].split("：")[1].split("{")[0])
                sk.append(ratio[r].split("：")[1].split("{")[0])
            if r == 5:
                if '总伤害' in ratio[r]:
                    round.append(ratio[r][4:].replace('\n', ''))
                    sk.append(ratio[r][4:].replace('\n', ''))
        sk.append(path)
        sk.append(line)
        Total.append(round)
        SkillData.append(sk)
        SkillSkill = list(set(SkillSkill).union(set(maybeskill)))
    # Total.append(["", "", "", "", "", "", "", "", "", "", "", "", "", ""])
    # for t in Total:
    #     print(t)
    return Total


def findtxt(path):
    txts = {}
    files = os.listdir(path)
    # print(files)
    temptxt = []
    for f in files:
        if os.path.isdir(path + '\\' + f):
            root = path + '\\' + f
            stip = os.listdir(path + '\\' + f)
            stip.sort(key=lambda x: float(x.split('.')[0].replace("-", ".")[:4]))
            txts[root] = stip
        if os.path.isfile(path + '\\' + f):
            temptxt.append(f)
    txts[path] = temptxt
    return txts


def writetoxlsx():
    global Total, SkillData, SkillSkill
    paths = findtxt(BattlPath)
    for r, ps in paths.items():
        for p in ps:
            getlog(r+'\\'+p)
    wb = load_workbook(ToolPath)
    ws = wb.create_sheet()
    for i in Total:
        if i != []:
            ws.append(i)
    sks = wb.create_sheet()
    for i in range(1, len(SkillData)):
        if len(SkillData[i]) > 3:
            totalHurt = round(float(SkillData[i][0]) * float(SkillData[i][1]) * float(SkillData[i][4]) *
                              float(SkillData[i][6]) * float(SkillData[i][17]) * float(SkillData[i][18]) -
                              float(SkillData[i][29]))
            ratiox = round((1 + float(SkillData[i][2]) + float(SkillData[i][3])), 2)
            ratioc = round((1 + float(SkillData[i][7])) * (1 + float(SkillData[i][8])) * (1 + float(SkillData[i][9])) * \
                     (1 + float(SkillData[i][10])) * (1 + float(SkillData[i][11])) * (1 + float(SkillData[i][12])) * \
                     (1 + float(SkillData[i][13])) * (1 + float(SkillData[i][14])) * (1 + float(SkillData[i][15])) *
                           (1 + float(SkillData[i][16])), 2)
            ratiod = round((1 - float(SkillData[i][19])) * (1 - float(SkillData[i][20])) * (1 - float(SkillData[i][21])) * \
                     (1 - float(SkillData[i][22])) * (1 - float(SkillData[i][23])) * (1 - float(SkillData[i][24])) * \
                     (1 - float(SkillData[i][25])) * (1 - float(SkillData[i][26])) * (1 - float(SkillData[i][27])) *
                           (1 - float(SkillData[i][28])), 2)
            if Total[i][3] not in ["Invincible", "Block", "Immune", "SecKill"]:
                if ratiox != float(SkillData[i][1]):
                    # sks.append([SkillData[i][31]])
                    # sks.append([SkillData[i][32]])
                    sks.append(["×××××× {0}、算出来的总X系数({1})与log({2})对不上!!!!!".format(i, ratiox, float(SkillData[i][1]))])
                if float(SkillData[i][4]) != float(SkillData[i][5].split("；")[0]):
                    # sks.append([SkillData[i][31]])
                    # sks.append([SkillData[i][32]])
                    sks.append(["×××××× {0}、算出来的总B系数({1})与log({2})对不上!!!!!".format(i, float(SkillData[i][4]),
                                                                             float(SkillData[i][5].split("；")[0]))])
                if ratioc != float(SkillData[i][6]):
                    # sks.append([SkillData[i][31]])
                    # sks.append([SkillData[i][32]])
                    sks.append(["×××××× {0}、算出来的总C系数({1})与log({2})对不上!!!!!".format(i, ratioc, float(SkillData[i][6]))])
                if ratiod != float(SkillData[i][18]):
                    sks.append([SkillData[i][31]])
                    sks.append([SkillData[i][32]])
                    sks.append(["×××××× {0}、算出来的总D系数({1})与log({2})对不上!!!!!".format(i, ratiod, float(SkillData[i][18]))])
            else:
                if '0' != SkillData[i][1]:
                    sks.append(["×××××× {0}、算出来的总X系数({1})与log({2})对不上!!!!!".format(i, '0', SkillData[i][1])])
                if SkillData[i][4] != "0":
                    sks.append(["×××××× {0}、算出来的总B系数({1})与log({2})对不上!!!!!".format(i, SkillData[i][4], "0")])
                if "0" != SkillData[i][6]:
                    sks.append(["×××××× {0}、算出来的总C系数({1})与log({2})对不上!!!!!".format(i, "0", SkillData[i][6])])
                if "0" != SkillData[i][18]:
                    sks.append(["×××××× {0}、算出来的总D系数({1})与log({2})对不上!!!!!".format(i, "0", SkillData[i][18])])
            if Total[i][2] in ["海德拉之小头", "海德拉之大头"]:
                totalHurt = 1
            if totalHurt != int(SkillData[i][30]):
                sks.append([SkillData[i][31]])
                sks.append([SkillData[i][32]])
                sks.append(["×××××× {0}、算出来的总伤害({1})与log({2})对不上!!!!!".format(i, totalHurt, int(SkillData[i][30]))])
    sss = wb.create_sheet()
    for i in SkillSkill:
        sss.append([i])
    wb.save(ToolPath)
    print("excel 导入完毕！！！！！！")


def getSkillids():
    book = xlrd.open_workbook(SkillXls)
    table = book.sheet_by_index(0)
    ids = table.col_values(0)
    dec = table.col_values(1)


if __name__ == "__main__":
    writetoxlsx()
    # checkFormula()